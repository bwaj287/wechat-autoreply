from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from erge_gateway.schemas import IngestResult


def ingest_xlsx(path: Path) -> IngestResult:
    workbook = load_workbook(filename=str(path), data_only=False, read_only=True)
    text_chunks: list[str] = []
    formula_count = 0
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names[:10]:
        sheet = workbook[sheet_name]
        rows: list[list[str]] = []
        for row in sheet.iter_rows(max_row=6, values_only=False):
            rendered: list[str] = []
            has_any = False
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                if value not in (None, ""):
                    has_any = True
                rendered.append("" if value is None else str(value))
            if has_any:
                rows.append(rendered)
        if rows:
            sample = "\n".join(" | ".join(r) for r in rows[:5])
            text_chunks.append(f"Sheet: {sheet_name}\n{sample}")
    return IngestResult(
        kind="xlsx",
        source_path=str(path),
        text_chunks=text_chunks,
        metadata={
            "sheet_names": sheet_names,
            "formula_count": formula_count,
            "extraction_mode": "text_native",
        },
        routing_hint="xlsx_structured",
    )
